"""
API עבור מערכת דוחות (שלב 11): תצוגה, ייצוא CSV, ייצוא Excel, והדפסה/PDF.
"""
import csv
import io

from flask import Blueprint, jsonify, request, render_template, Response
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services import reports_service as svc
from app.services.reports_service import ReportServiceError

reports_bp = Blueprint('reports', __name__)


@reports_bp.errorhandler(ReportServiceError)
def _handle_error(err):
    return jsonify({"error": str(err)}), 400


@reports_bp.route('/reports')
def reports_page():
    return render_template('reports.html', active_nav='reports')


@reports_bp.route('/reports/<report_key>/print')
def report_print_page(report_key):
    report = svc.get_report(report_key)
    return render_template('report_print.html', report=report)


@reports_bp.route('/api/reports', methods=['GET'])
def api_list_reports():
    return jsonify(svc.list_report_types())


@reports_bp.route('/api/reports/<report_key>', methods=['GET'])
def api_get_report(report_key):
    report = svc.get_report(report_key)
    return jsonify(report)


@reports_bp.route('/api/reports/<report_key>/csv', methods=['GET'])
def api_export_csv(report_key):
    report = svc.get_report(report_key)
    buf = io.StringIO()
    buf.write('\ufeff')  # BOM כדי שאקסל יזהה נכון עברית ב-UTF-8
    writer = csv.writer(buf)
    writer.writerow(report["headers"])
    writer.writerows(report["rows"])
    return Response(
        buf.getvalue(), mimetype='text/csv',
        headers={"Content-Disposition": f"attachment; filename={report_key}.csv"},
    )


@reports_bp.route('/api/reports/<report_key>/excel', methods=['GET'])
def api_export_excel(report_key):
    report = svc.get_report(report_key)
    wb = Workbook()
    ws = wb.active
    ws.title = report["label"][:31]
    ws.sheet_view.rightToLeft = True
    ws.append(report["headers"])
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for row in report["rows"]:
        ws.append(row)
    for i, header in enumerate(report["headers"], start=1):
        width = max(12, len(str(header)) + 4)
        ws.column_dimensions[get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename={report_key}.xlsx"},
    )
